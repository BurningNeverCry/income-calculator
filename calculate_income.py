#!/usr/bin/env python3
"""
境外证券收益计算器 - 加权平均成本法
根据富途牛牛年度账单Excel计算股息收入和财产转让所得。
输出JSON格式供报告生成使用。
"""
import sys
import os
import glob
import re
import json
import argparse
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("错误：需要安装 openpyxl: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# ──────────────────────────────────────────────
# 加权平均成本追踪器
# ──────────────────────────────────────────────
class AvgCostTracker:
    """使用加权平均成本法追踪持仓。"""

    def __init__(self):
        self.holdings = {}  # {symbol: {'qty', 'total_cost', 'avg_cost', 'currency'}}

    def _normalize_symbol(self, symbol):
        """处理代码转换，如 DIDI -> DIDIY"""
        s = symbol.strip().upper()
        if s == 'DIDI':
            return 'DIDIY'
        return s

    def buy(self, symbol, qty, amount, fee, currency):
        """
        买入：更新持仓均价。
        amount: 成交金额绝对值
        fee: 手续费绝对值
        """
        symbol = self._normalize_symbol(symbol)
        total_cost_this = amount + fee

        if symbol not in self.holdings:
            self.holdings[symbol] = {'qty': 0, 'total_cost': 0.0, 'avg_cost': 0.0, 'currency': currency}

        h = self.holdings[symbol]
        h['qty'] += qty
        h['total_cost'] += total_cost_this
        h['avg_cost'] = h['total_cost'] / h['qty'] if h['qty'] > 0 else 0
        h['currency'] = currency

    def sell(self, symbol, qty, amount, fee, currency):
        """
        卖出：以当前均价计算成本基础，均价不变。
        返回 (avg_cost, cost_basis, gain)
        """
        symbol = self._normalize_symbol(symbol)
        net_proceeds = amount - fee

        if symbol not in self.holdings or self.holdings[symbol]['qty'] <= 0:
            # 无持仓信息，成本为0
            return 0, 0, net_proceeds

        h = self.holdings[symbol]
        avg_cost = h['avg_cost']
        cost_basis = qty * avg_cost
        gain = net_proceeds - cost_basis

        h['qty'] -= qty
        h['total_cost'] = h['qty'] * avg_cost  # 均价不变
        if h['qty'] <= 0:
            h['qty'] = 0
            h['total_cost'] = 0
            h['avg_cost'] = 0

        return avg_cost, cost_basis, gain

    def get_avg_cost(self, symbol):
        symbol = self._normalize_symbol(symbol)
        if symbol in self.holdings:
            return self.holdings[symbol]['avg_cost']
        return 0


# ──────────────────────────────────────────────
# 文件扫描
# ──────────────────────────────────────────────
def scan_year_files(folder_path):
    """扫描文件夹，返回按年份排序的 {year: filepath}"""
    xlsx_files = glob.glob(os.path.join(folder_path, '*.xlsx'))
    year_files = {}
    for fpath in sorted(xlsx_files):
        fname = os.path.basename(fpath)
        matches = re.findall(r'(20[2-3]\d)', fname)
        for year_str in matches:
            year = int(year_str)
            if year not in year_files:
                year_files[year] = fpath
    return dict(sorted(year_files.items()))


# ──────────────────────────────────────────────
# 交易流水读取
# ──────────────────────────────────────────────
def read_trades(filepath):
    """读取证券-交易流水，返回交易列表。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if '证券-交易流水' not in wb.sheetnames:
        wb.close()
        return []

    ws = wb['证券-交易流水']
    headers = [cell.value for cell in ws[1]]
    trades = []

    for r in range(2, ws.max_row + 1):
        rd = {}
        for c, h in enumerate(headers, 1):
            rd[h] = ws.cell(row=r, column=c).value

        if not rd.get('成交时间'):
            continue

        date_str = str(rd['成交时间'])[:10]
        code_name = str(rd.get('代码名称') or '').strip()
        category = str(rd.get('品类') or '').strip()
        direction = str(rd.get('方向') or '').strip()
        market = str(rd.get('交易所/市场') or '').strip()
        currency = str(rd.get('币种') or '').strip()
        qty = abs(float(rd.get('数量/面值') or 0))
        price = abs(float(rd.get('价格') or 0))
        amount = abs(float(rd.get('成交金额') or 0))
        fee = abs(float(rd.get('总费用') or 0))

        trades.append({
            'date': date_str,
            'code': code_name,
            'category': category,
            'direction': direction,
            'market': market,
            'currency': currency,
            'qty': qty,
            'price': price,
            'amount': amount,
            'fee': fee,
        })

    wb.close()
    return trades


# ──────────────────────────────────────────────
# 股息读取
# ──────────────────────────────────────────────
def _parse_stock_from_remark(remark):
    """从备注中解析股票代码/名称。"""
    import re
    if not remark:
        return ''
    remark = str(remark)
    # 匹配 <SEHK 1398 ICBC> 格式
    m = re.search(r'<SEHK\s+(\d+)\s+([^>]+)>', remark)
    if m:
        return f"{m.group(1)} {m.group(2).strip()}"
    # 匹配 BABA xxx SHARES DIVIDENDS 格式
    m = re.search(r'^([A-Z]+)\s+[\d.]+\s+SHARES', remark)
    if m:
        return m.group(1)
    # 匹配 ADR FEE 格式
    m = re.search(r'^([A-Z]+)\s+[\d.]+\s+SHARES\s+ADR', remark)
    if m:
        return m.group(1)
    return ''


def read_dividends(filepath):
    """
    读取证券-资金进出，提取股息记录。
    列结构：日期 | 账户名称 | 账户号码 | 类型 | 方向 | 币种 | 变动金额 | 备注
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if '证券-资金进出' not in wb.sheetnames:
        wb.close()
        return []

    ws = wb['证券-资金进出']
    headers = [cell.value for cell in ws[1]]
    dividends = []

    # 手续费关键词（备注中出现即为费用）
    fee_keywords = ['handling charge', 'scrip charge', 'dividend fee']
    # 股息关键词（备注中出现即为股息收入）
    dividend_keywords = ['I/D', 'F/D', 'DIVIDEND']

    for r in range(2, ws.max_row + 1):
        rd = {}
        for c, h in enumerate(headers, 1):
            rd[h] = ws.cell(row=r, column=c).value

        # 只处理"公司行动"或"资金进出"类型（不同年份格式不同）
        type_val = str(rd.get('类型') or '').strip()
        if '公司行动' not in type_val and '资金进出' not in type_val:
            continue

        date_str = str(rd.get('日期') or '')[:10]
        direction = str(rd.get('方向') or '').strip()  # In or Out
        currency = str(rd.get('币种') or '').strip()
        amount = float(rd.get('变动金额') or 0)
        remark = str(rd.get('备注') or '').strip()
        remark_upper = remark.upper()

        # 备注中必须含有股息关键词或手续费关键词才继续处理
        has_dividend_keyword = any(kw.upper() in remark_upper for kw in dividend_keywords)
        has_fee_keyword = any(kw.upper() in remark_upper for kw in fee_keywords)
        if not has_dividend_keyword and not has_fee_keyword:
            continue

        # 判断是否为手续费
        is_fee = has_fee_keyword

        # 判断是否为股息收入（Direction=In）
        is_dividend = (direction == 'In' and has_dividend_keyword)

        if not is_dividend and not is_fee:
            continue

        # 从备注解析股票名称
        stock_name = _parse_stock_from_remark(remark)

        dividends.append({
            'date': date_str,
            'stock': stock_name,
            'currency': currency,
            'amount': round(amount, 2),
            'is_fee': is_fee,
        })

    wb.close()
    return dividends


# ──────────────────────────────────────────────
# 主计算逻辑
# ──────────────────────────────────────────────
def calculate_income(folder_path, target_years):
    """
    主计算函数。
    返回结果字典：{year: {dividends: [...], trades: [...], options: [...]}}
    """
    year_files = scan_year_files(folder_path)
    if not year_files:
        print("错误：未找到年度账单文件", file=sys.stderr)
        sys.exit(1)

    tracker = AvgCostTracker()
    results = {y: {'dividends': [], 'trades': [], 'options': [], 'summary': {}} for y in target_years}

    # ── 处理所有年份的交易流水 ──
    all_trades = []
    for year, fpath in year_files.items():
        trades = read_trades(fpath)
        for t in trades:
            t['year'] = year
        all_trades.extend(trades)

    # 按日期排序
    all_trades.sort(key=lambda x: x['date'])

    # 逐笔处理
    for t in all_trades:
        year = int(t['date'][:4])
        direction = t['direction']
        is_option = t['category'] == '期权'

        if '买' in direction:
            if is_option and year in target_years:
                # 记录期权买入（后续可能到期作废）
                pass
            tracker.buy(t['code'], t['qty'], t['amount'], t['fee'], t['currency'])

        elif '卖' in direction:
            avg_cost, cost_basis, gain = tracker.sell(t['code'], t['qty'], t['amount'], t['fee'], t['currency'])

            if year in target_years:
                if is_option:
                    results[year]['options'].append({
                        'date': t['date'],
                        'contract': t['code'],
                        'action': f"卖出",
                        'gain': round(gain, 2),
                        'currency': t['currency'],
                    })
                else:
                    results[year]['trades'].append({
                        'date': t['date'],
                        'stock': t['code'],
                        'qty': int(t['qty']),
                        'price': round(t['price'], 4),
                        'proceeds': round(t['amount'] - t['fee'], 2),
                        'avg_cost': round(avg_cost, 4),
                        'cost_basis': round(cost_basis, 2),
                        'gain': round(gain, 2),
                        'currency': t['currency'],
                    })

    # ── 处理期权到期作废（品类=期权，有买入但无对应卖出的，在到期日计为损失）──
    # 通过检查持仓中还有期权残留来处理
    # 简化处理：期权如果买入后持仓清零，说明到期作废
    option_buys = defaultdict(list)
    option_sells = set()
    for t in all_trades:
        if t['category'] == '期权':
            if '买' in t['direction']:
                option_buys[t['code']].append(t)
            elif '卖' in t['direction']:
                option_sells.add(t['code'])

    for code, buys in option_buys.items():
        if code not in option_sells:
            # 到期作废
            for b in buys:
                year = int(b['date'][:4])
                # 到期作废的损失应该在到期年份计，这里简化为买入年份
                if year in target_years:
                    loss = -(b['amount'] + b['fee'])
                    results[year]['options'].append({
                        'date': b['date'],
                        'contract': code,
                        'action': f"到期作废 {int(b['qty'])} 张",
                        'gain': round(loss, 2),
                        'currency': b['currency'],
                    })

    # ── 处理股息 ──
    for year, fpath in year_files.items():
        if year not in target_years:
            continue
        divs = read_dividends(fpath)
        for d in divs:
            if not d['is_fee']:
                results[year]['dividends'].append({
                    'date': d['date'],
                    'stock': d['stock'],
                    'currency': d['currency'],
                    'amount': round(d['amount'], 2),
                })
            else:
                # 记录手续费（负值）
                results[year]['dividends'].append({
                    'date': d['date'],
                    'stock': d['stock'] + ' (手续费)',
                    'currency': d['currency'],
                    'amount': round(d['amount'], 2),
                    'is_fee': True,
                })

    # ── 计算各年汇总 ──
    for year in target_years:
        r = results[year]
        # 股息汇总（按币种，包含手续费减扣）
        div_by_currency = defaultdict(float)
        for d in r['dividends']:
            div_by_currency[d['currency']] += d['amount']

        # 交易汇总（按币种）
        trade_gain_by_currency = defaultdict(float)
        for t in r['trades']:
            trade_gain_by_currency[t['currency']] += t['gain']

        # 期权汇总（按币种）
        option_gain_by_currency = defaultdict(float)
        for o in r['options']:
            option_gain_by_currency[o['currency']] += o['gain']

        r['summary'] = {
            'dividends': dict(div_by_currency),
            'capital_gains_securities': dict(trade_gain_by_currency),
            'capital_gains_options': dict(option_gain_by_currency),
        }

    return results


# ──────────────────────────────────────────────
# CLI 入口
# ──────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='境外证券收益计算器')
    parser.add_argument('folder', help='年度账单文件夹路径')
    parser.add_argument('--target-years', required=True, help='目标年份，逗号分隔，如 2023,2024,2025')
    parser.add_argument('--output', default='-', help='输出文件路径，默认stdout')

    args = parser.parse_args()
    target_years = [int(y.strip()) for y in args.target_years.split(',')]

    results = calculate_income(args.folder, target_years)

    output = json.dumps(results, ensure_ascii=False, indent=2)

    if args.output == '-':
        print(output)
    else:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(output)
        print(f"结果已保存到: {args.output}", file=sys.stderr)


if __name__ == '__main__':
    main()
